import pathlib
from ..workflow_obj import workflow_obj
import tarfile
import os 
import datetime
from workflow.logger import Script_Logger
import subprocess
from pathlib import Path
from openpyxl import load_workbook
import csv





class gisaid_submit_obj(workflow_obj):

    def __init__(self):
        self.id = "gisaid_submit"
        #elf.current_gis_path = "/home/ssh_user/WGS_Sequencing_COVID/GISAID/"
        self.log= Script_Logger("GISAID_Submit")
        self.log.start_log("Initialization of GISAID SUBMIT sucessful")


    def get_json(self):
        super().get_json(self.id) #what does this need besides    
        self.log.write_log("get_json",self.id+" passed")
        

    def submit(self,run_id):
        #TODO
        #first need to check if token is valid - will prob need to do this last
        #email CLIS support 
        #if token is expired get user to re-enter password micheal only RN

       
        date = str(datetime.datetime.strptime(run_id[7:17], "%Y-%m-%d").strftime("%m%d%y"))
        self.current_gis_path += date

        #look for upload files
        for file in os.listdir(self.current_gis_path):
            if file.endswith(".fasta"):
                f_num = file[-8:-6]
                #file contains correct fasta file now find coressponding upload file
              
                date_2 = [date[i:i+2] for i in range(0,len(date),2)]
              
                #create metadata upload 
                bulk_metadata = self.create_bulk_upload_template(date_2,self.current_gis_path+"/20"+date_2[2]+date_2[0]+date_2[1]+f_num+"_sql.xlsx",self.current_gis_path,f_num)


       
                if os.path.exists(self.current_gis_path+"/"+bulk_metadata):
                    #if bulk file is there
                 
                    exec_cmd="cli2 upload --token "+self.token_path+" --metadata "+self.current_gis_path+"/"+bulk_metadata+" --fasta "+self.current_gis_path+"/"+file+" --frameshift catch_all --log "+self.log_path+" --failed "+self.current_gis_path+"/"+date+"_FAILED_GISAID.csv"
                    subprocess.run(exec_cmd,shell=True)
                    #self.log.write_log("Submit","Output from command \n"+command_ouput)
                  



    def create_new_token(self):
        #"cli2 authenticate --token"+path_to_token+"GIS_TOKEN --username "+self.user+" --client_id "+self.user_client_ID
        token_cmd = "cli2 authenticate --token"+self.token_path+"/GIS_TOKEN --username "+self.user+" --client_id "+self.user_client_ID
        command_ouput= subprocess.run(token_cmd,capture_output=True,shell=True)
        self.log.write_log("creat_new_token_ouput", command_ouput)


    def create_bulk_upload_template(self,date, sql_file, output_path,file_number):#going to need to write this
        
        #first read in sql metadata
        workbook = load_workbook(filename=sql_file, read_only=True)
        sheet=workbook.active
        #create output file
        gis_metadata = open(output_path+"/20210222_EpiCoV_BulkUpload_Template_"+date[0]+"_"+date[1]+"_20"+date[2]+file_number+".csv", 'w+')

        # create the csv writer
        writer = csv.writer(gis_metadata)

        header=["submitter","fn","covv_virus_name","covv_type","covv_passage","covv_collection_date","covv_location","covv_add_location","covv_host","covv_add_host_info","covv_sampling_strategy","covv_gender","covv_patient_age","covv_patient_status","covv_specimen","covv_outbreak","covv_last_vaccinated","covv_treatment","covv_seq_technology","covv_assembly_method","covv_coverage","covv_orig_lab","covv_orig_lab_addr","covv_provider_sample_id","covv_subm_lab","covv_subm_lab_addr","covv_subm_sample_id","covv_authors","covv_comment","comment_type"]
        writer.writerow(header)
        
        #then loop through sql while wirteing into excel file
        for value in sheet.iter_rows(min_row=2,min_col=2,max_col=31,max_row=sheet.max_row, values_only=True):
            writer.writerow(value)
        #close openpyxl
        workbook.close()
        #close csv
        gis_metadata.close()

        return "20210222_EpiCoV_BulkUpload_Template_"+date[0]+"_"+date[1]+"_20"+date[2]+file_number+".csv"
        
